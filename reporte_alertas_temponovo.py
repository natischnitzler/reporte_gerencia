"""
Reporte Automático de Alertas - Temponovo
Envío: Lunes y jueves
Contenido: Descuentos altos | Cobranza vencida | Pedidos atrasados
Formato: Email HTML + 3 Excel adjuntos
"""

import xmlrpc.client
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, date, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════
# CONFIGURACIÓN
# Las credenciales se leen desde variables de entorno (GitHub Secrets)
# Para pruebas locales puedes setearlas en tu terminal:
#   export ODOO_URL="https://tu-odoo.temponovo.cl"
#   export ODOO_PASSWORD="mi_clave"  etc.
# ══════════════════════════════════════════════
import os

ODOO_URL      = os.environ.get("ODOO_URL", "")
ODOO_DB       = os.environ.get("ODOO_DB", "temponovo")
ODOO_USER     = os.environ.get("ODOO_USER", "")
ODOO_PASSWORD = os.environ.get("ODOO_PASSWORD", "")

SMTP_HOST     = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER     = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")

DESTINATARIOS = [
    "daniel@temponovo.cl",
    "natalia@temponovo.cl",
]

# ── Umbrales ──
DESC_AMARILLO   = 30.0   # % → alerta amarilla
DESC_ROJO       = 50.0   # % → alerta roja
COBR_DIAS       = 30     # días vencido mínimo para cobranza
DIAS_COTIZACION = 3      # días en cotización sin confirmar
DIAS_SIN_PICK   = 3      # días confirmado sin TN/PICK realizado
DIAS_SIN_OUT    = 3      # días con PICK pero sin TN/OUT realizado


# ══════════════════════════════════════════════
# CONEXIÓN ODOO
# ══════════════════════════════════════════════
def conectar_odoo():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
    if not uid:
        raise Exception("Autenticación fallida. Verifica credenciales Odoo.")
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    return uid, models

def buscar(models, uid, modelo, dominio, campos, limite=1000):
    return models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        modelo, 'search_read',
        [dominio],
        {'fields': campos, 'limit': limite}
    )


# ══════════════════════════════════════════════
# 1. DESCUENTOS > 30%
# ══════════════════════════════════════════════
def get_descuentos(models, uid):
    """
    Líneas de pedidos confirmados y facturas con descuento > 30%.
    Resumen email: Cliente | N° Pedido | % Descuento
    Excel detalle: + Código | Producto | Precio | Cantidad | Subtotal
    """
    lineas_pedido = buscar(models, uid,
        'sale.order.line',
        [
            ['discount', '>', DESC_AMARILLO],
            ['order_id.state', 'in', ['sale', 'done']],
        ],
        ['order_id', 'product_id', 'discount', 'price_unit', 'product_uom_qty', 'price_subtotal']
    )

    lineas_factura = buscar(models, uid,
        'account.move.line',
        [
            ['discount', '>', DESC_AMARILLO],
            ['move_id.move_type', '=', 'out_invoice'],
            ['move_id.state', '=', 'posted'],
            ['exclude_from_invoice_tab', '=', False],
        ],
        ['move_id', 'partner_id', 'product_id', 'discount', 'price_unit', 'quantity', 'price_subtotal']
    )

    resumen = []
    detalle = []

    for l in lineas_pedido:
        nivel   = '🔴 Rojo' if l['discount'] >= DESC_ROJO else '🟡 Amarillo'
        cliente = l['order_id'][1].split(' - ')[0] if l['order_id'] else ''
        pedido  = l['order_id'][1] if l['order_id'] else ''
        prod    = l['product_id'][1] if l['product_id'] else ''
        codigo  = prod.split(']')[0].replace('[','').strip() if ']' in prod else ''
        nombre  = prod.split('] ')[-1] if ']' in prod else prod

        resumen.append({'Cliente': cliente, 'N° Pedido': pedido,
                        'Descuento': l['discount'], 'Nivel': nivel})
        detalle.append({
            'Origen': 'Pedido', 'Cliente': cliente, 'N° Pedido': pedido,
            'Código': codigo, 'Producto': nombre,
            'Precio Unit': l['price_unit'], 'Descuento %': l['discount'],
            'Cantidad': l['product_uom_qty'], 'Subtotal': l['price_subtotal'],
            'Nivel': nivel,
        })

    for l in lineas_factura:
        nivel   = '🔴 Rojo' if l['discount'] >= DESC_ROJO else '🟡 Amarillo'
        cliente = l['partner_id'][1] if l['partner_id'] else ''
        factura = l['move_id'][1] if l['move_id'] else ''
        prod    = l['product_id'][1] if l['product_id'] else ''
        codigo  = prod.split(']')[0].replace('[','').strip() if ']' in prod else ''
        nombre  = prod.split('] ')[-1] if ']' in prod else prod

        resumen.append({'Cliente': cliente, 'N° Pedido': factura,
                        'Descuento': l['discount'], 'Nivel': nivel})
        detalle.append({
            'Origen': 'Factura', 'Cliente': cliente, 'N° Pedido': factura,
            'Código': codigo, 'Producto': nombre,
            'Precio Unit': l['price_unit'], 'Descuento %': l['discount'],
            'Cantidad': l['quantity'], 'Subtotal': l['price_subtotal'],
            'Nivel': nivel,
        })

    resumen.sort(key=lambda x: x['Descuento'], reverse=True)
    detalle.sort(key=lambda x: x['Descuento %'], reverse=True)
    return resumen, detalle


# ══════════════════════════════════════════════
# 2. COBRANZA VENCIDA +30 DÍAS
# ══════════════════════════════════════════════
def get_cobranza(models, uid):
    """
    Replica vista Odoo: Cliente · Vendedor · Ciudad | A la fecha | 1-30 | Vencido >30 | Total
    Solo clientes con algo vencido >30 días.
    """
    hoy = date.today()

    facturas = buscar(models, uid,
        'account.move',
        [
            ['move_type', '=', 'out_invoice'],
            ['payment_state', 'in', ['not_paid', 'partial']],
            ['state', '=', 'posted'],
        ],
        ['name', 'partner_id', 'invoice_date_due', 'amount_residual', 'invoice_user_id']
    )

    clientes = {}
    for f in facturas:
        pid = f['partner_id'][0] if f['partner_id'] else 0
        if pid not in clientes:
            clientes[pid] = {
                'Cliente': f['partner_id'][1] if f['partner_id'] else '',
                'Vendedor': f['invoice_user_id'][1] if f['invoice_user_id'] else 'Sin vendedor',
                'Ciudad': '',
                'A la fecha': 0.0, '1-30': 0.0, 'Vencido >30': 0.0, 'Total': 0.0,
                'facturas': [],
            }

        venc_str = f.get('invoice_date_due') or ''
        monto    = f['amount_residual'] or 0.0

        if venc_str:
            venc = datetime.strptime(venc_str, '%Y-%m-%d').date()
            dias = (hoy - venc).days
            if dias <= 0:
                clientes[pid]['A la fecha'] += monto
            elif dias <= 30:
                clientes[pid]['1-30'] += monto
            else:
                clientes[pid]['Vencido >30'] += monto
        else:
            clientes[pid]['A la fecha'] += monto

        clientes[pid]['Total'] += monto
        dias_v = (hoy - datetime.strptime(venc_str, '%Y-%m-%d').date()).days if venc_str else 0
        clientes[pid]['facturas'].append({
            'Factura': f['name'], 'Fecha Venc.': venc_str,
            'Días Vencido': dias_v, 'Monto Pendiente': monto,
        })

    # Enriquecer ciudad
    if clientes:
        partners = buscar(models, uid, 'res.partner',
            [['id', 'in', list(clientes.keys())]], ['id', 'city'])
        for p in partners:
            if p['id'] in clientes:
                clientes[p['id']]['Ciudad'] = p.get('city') or ''

    resumen = [v for v in clientes.values() if v['Vencido >30'] > 0]
    resumen.sort(key=lambda x: x['Vencido >30'], reverse=True)

    detalle = []
    for cli in resumen:
        for fac in cli['facturas']:
            detalle.append({
                'Cliente': cli['Cliente'], 'Vendedor': cli['Vendedor'],
                'Ciudad': cli['Ciudad'], 'Factura': fac['Factura'],
                'Fecha Venc.': fac['Fecha Venc.'], 'Días Vencido': fac['Días Vencido'],
                'Monto Pendiente': fac['Monto Pendiente'],
            })

    return resumen, detalle


# ══════════════════════════════════════════════
# 3. PEDIDOS ATRASADOS
# ══════════════════════════════════════════════
def get_pedidos_atrasados(models, uid):
    """
    3 estados de alerta:
    - En cotización   : draft > 3 días
    - No pickeado     : confirmado, ningún TN/PICK realizado > 3 días
    - No en bulto     : TN/PICK realizado pero sin TN/OUT realizado > 3 días
    """
    hoy = date.today()
    alertas = []

    # 1. Cotizaciones sin confirmar
    cotizaciones = buscar(models, uid, 'sale.order',
        [
            ['state', '=', 'draft'],
            ['date_order', '<', (hoy - timedelta(days=DIAS_COTIZACION)).strftime('%Y-%m-%d %H:%M:%S')],
        ],
        ['name', 'partner_id', 'date_order', 'amount_total', 'user_id']
    )
    for p in cotizaciones:
        dias = (hoy - datetime.strptime(p['date_order'][:10], '%Y-%m-%d').date()).days
        alertas.append({
            'N° Pedido': p['name'],
            'Cliente':   p['partner_id'][1] if p['partner_id'] else '',
            'Vendedor':  p['user_id'][1] if p['user_id'] else '',
            'Estado':    'En cotización',
            'Umbral':    f'+{DIAS_COTIZACION} días',
            'Días':      dias,
            'Monto':     p['amount_total'],
        })

    # 2 y 3. Pedidos confirmados → revisar pickings
    confirmados = buscar(models, uid, 'sale.order',
        [
            ['state', '=', 'sale'],
            ['date_order', '<', (hoy - timedelta(days=DIAS_SIN_PICK)).strftime('%Y-%m-%d %H:%M:%S')],
        ],
        ['name', 'partner_id', 'date_order', 'amount_total', 'user_id', 'picking_ids']
    )

    for p in confirmados:
        picking_ids = p.get('picking_ids', [])
        dias = (hoy - datetime.strptime(p['date_order'][:10], '%Y-%m-%d').date()).days
        base = {
            'N° Pedido': p['name'],
            'Cliente':   p['partner_id'][1] if p['partner_id'] else '',
            'Vendedor':  p['user_id'][1] if p['user_id'] else '',
            'Días':      dias,
            'Monto':     p['amount_total'],
        }

        if not picking_ids:
            alertas.append({**base, 'Estado': 'No pickeado', 'Umbral': f'+{DIAS_SIN_PICK} días'})
            continue

        pickings = buscar(models, uid, 'stock.picking',
            [['id', 'in', picking_ids]],
            ['name', 'state']
        )

        picks = [pk for pk in pickings if 'PICK' in (pk.get('name') or '')]
        outs  = [pk for pk in pickings if 'OUT'  in (pk.get('name') or '')]

        pick_realizado = any(pk['state'] == 'done' for pk in picks)
        out_realizado  = any(pk['state'] == 'done' for pk in outs)

        if not pick_realizado:
            alertas.append({**base, 'Estado': 'No pickeado', 'Umbral': f'+{DIAS_SIN_PICK} días'})
        elif not out_realizado and dias >= DIAS_SIN_OUT:
            alertas.append({**base, 'Estado': 'No en bulto', 'Umbral': f'+{DIAS_SIN_OUT} días'})

    alertas.sort(key=lambda x: x['Días'], reverse=True)
    return alertas


# ══════════════════════════════════════════════
# HELPERS EXCEL
# ══════════════════════════════════════════════
AZUL    = '1B3A6B'
BLANCO  = 'FFFFFF'
ROJO_BG = 'FFEBEE'
AMA_BG  = 'FFFDE7'
GRIS_BG = 'F5F5F5'

def _h(cell, bg=AZUL, fg=BLANCO):
    cell.font      = Font(bold=True, color=fg, name='Arial', size=10)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = Border(bottom=Side(style='thin', color='CCCCCC'),
                            right =Side(style='thin', color='CCCCCC'))

def _d(cell, bg=BLANCO, bold=False):
    cell.font      = Font(name='Arial', size=9, bold=bold)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(vertical='center')
    cell.border    = Border(bottom=Side(style='thin', color='E0E0E0'),
                            right =Side(style='thin', color='E0E0E0'))

def _titulo(ws, texto, n_cols):
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    c = ws['A1']
    c.value     = texto
    c.font      = Font(bold=True, size=13, color=BLANCO, name='Arial')
    c.fill      = PatternFill('solid', start_color=AZUL)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

def _autowidth(ws, headers, datos):
    for i, h in enumerate(headers, 1):
        w = max(len(str(h)) + 4, max((min(len(str(r.get(h,'')))+2, 45) for r in datos), default=10))
        ws.column_dimensions[get_column_letter(i)].width = w

def _clp(cell):
    cell.number_format = '$#,##0;($#,##0);"-"'

def _pct(cell):
    cell.number_format = '0.0"%"'


# ══════════════════════════════════════════════
# EXCEL 1: DESCUENTOS
# ══════════════════════════════════════════════
def excel_descuentos(detalle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descuentos"
    ws.sheet_properties.tabColor = "D32F2F"

    headers = ['Origen','Cliente','N° Pedido','Código','Producto',
               'Precio Unit','Descuento %','Cantidad','Subtotal','Nivel']
    _titulo(ws, f'DESCUENTOS > {DESC_AMARILLO}%  —  Pedidos confirmados y Facturas', len(headers))

    for i, h in enumerate(headers, 1):
        _h(ws.cell(row=2, column=i, value=h))
    ws.row_dimensions[2].height = 24

    for ri, row in enumerate(detalle, 3):
        bg = ROJO_BG if '🔴' in row.get('Nivel','') else AMA_BG
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h,''))
            _d(c, bg)
            if h in ('Precio Unit','Subtotal'): _clp(c)
            elif h == 'Descuento %': _pct(c)

    if detalle:
        tr = len(detalle) + 3
        c = ws.cell(row=tr, column=len(headers), value=f'=SUM(I3:I{tr-1})')
        _d(c, GRIS_BG, bold=True); _clp(c)
        ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, name='Arial', size=9)

    _autowidth(ws, headers, detalle)
    ws.freeze_panes = 'A3'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════
# EXCEL 2: COBRANZA
# ══════════════════════════════════════════════
def excel_cobranza(detalle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobranza"
    ws.sheet_properties.tabColor = "FF6F00"

    headers = ['Cliente','Vendedor','Ciudad','Factura','Fecha Venc.','Días Vencido','Monto Pendiente']
    _titulo(ws, f'COBRANZA VENCIDA — Facturas con más de {COBR_DIAS} días', len(headers))

    for i, h in enumerate(headers, 1):
        _h(ws.cell(row=2, column=i, value=h))
    ws.row_dimensions[2].height = 24

    for ri, row in enumerate(detalle, 3):
        dias = row.get('Días Vencido', 0)
        bg   = ROJO_BG if dias > 90 else AMA_BG if dias > 60 else BLANCO
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h,''))
            _d(c, bg)
            if h == 'Monto Pendiente': _clp(c)

    if detalle:
        tr = len(detalle) + 3
        c = ws.cell(row=tr, column=7, value=f'=SUM(G3:G{tr-1})')
        _d(c, GRIS_BG, bold=True); _clp(c)
        ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, name='Arial', size=9)

    _autowidth(ws, headers, detalle)
    ws.freeze_panes = 'A3'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════
# EXCEL 3: PEDIDOS ATRASADOS
# ══════════════════════════════════════════════
def excel_pedidos(alertas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Atrasados"
    ws.sheet_properties.tabColor = "1565C0"

    headers = ['N° Pedido','Cliente','Vendedor','Estado','Umbral','Días','Monto']
    _titulo(ws, 'PEDIDOS ATRASADOS POR ETAPA', len(headers))

    for i, h in enumerate(headers, 1):
        _h(ws.cell(row=2, column=i, value=h))
    ws.row_dimensions[2].height = 24

    ESTADO_BG = {'En cotización': 'FFF9C4', 'No pickeado': 'FFE0B2', 'No en bulto': 'FFCDD2'}

    for ri, row in enumerate(alertas, 3):
        bg = ESTADO_BG.get(row.get('Estado',''), BLANCO)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h,''))
            _d(c, bg)
            if h == 'Monto': _clp(c)

    _autowidth(ws, headers, alertas)
    ws.freeze_panes = 'A3'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════
# EMAIL HTML
# ══════════════════════════════════════════════
def generar_html(desc_res, cobr_res, pedidos):
    hoy   = date.today().strftime('%d de %B de %Y').capitalize()
    total = len(desc_res) + len(cobr_res) + len(pedidos)
    total_vencido = sum(r['Vencido >30'] for r in cobr_res)

    def fmt(v):
        try: return f"$ {int(v):,}".replace(',','.')
        except: return str(v)

    def tabla_desc(rows, max_r=10):
        if not rows:
            return '<p style="color:#4caf50;font-style:italic;">✅ Sin alertas de descuentos</p>'
        t = '<table style="width:100%;border-collapse:collapse;font-size:12px;">'
        t += '<tr>' + ''.join(f'<th style="background:#1B3A6B;color:#fff;padding:8px 12px;text-align:left;">{h}</th>'
                               for h in ['Cliente','N° Pedido / Factura','Descuento','Nivel']) + '</tr>'
        for r in rows[:max_r]:
            bg = '#FFEBEE' if '🔴' in r['Nivel'] else '#FFFDE7'
            t += f'''<tr style="background:{bg};">
              <td style="padding:7px 12px;border-bottom:1px solid #eee;">{r["Cliente"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;">{r["N° Pedido"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;text-align:center;font-weight:bold;">{r["Descuento"]:.1f}%</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;text-align:center;">{r["Nivel"]}</td>
            </tr>'''
        if len(rows) > max_r:
            t += f'<tr><td colspan="4" style="text-align:center;padding:8px;color:#888;font-style:italic;">... y {len(rows)-max_r} más en el Excel adjunto</td></tr>'
        return t + '</table>'

    def tabla_cobr(rows, max_r=10):
        if not rows:
            return '<p style="color:#4caf50;font-style:italic;">✅ Sin facturas vencidas</p>'
        t = '<table style="width:100%;border-collapse:collapse;font-size:12px;">'
        t += '<tr>' + ''.join(f'<th style="background:#1B3A6B;color:#fff;padding:8px 12px;">{h}</th>'
                               for h in ['Cliente · Vendedor · Ciudad','A la fecha','1-30 días','Vencido &gt;30','Total']) + '</tr>'
        for r in rows[:max_r]:
            info = f"<strong>{r['Cliente']}</strong><br><small style='color:#888;'>{r['Vendedor']} · {r['Ciudad']}</small>"
            t += f'''<tr>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;">{info}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;">{fmt(r["A la fecha"]) if r["A la fecha"] else "—"}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;">{fmt(r["1-30"]) if r["1-30"] else "—"}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;font-weight:bold;color:#c62828;">{fmt(r["Vencido >30"])}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;font-weight:bold;">{fmt(r["Total"])}</td>
            </tr>'''
        if len(rows) > max_r:
            t += f'<tr><td colspan="5" style="text-align:center;padding:8px;color:#888;font-style:italic;">... y {len(rows)-max_r} clientes más en el Excel adjunto</td></tr>'
        return t + '</table>'

    def tabla_ped(rows, max_r=10):
        if not rows:
            return '<p style="color:#4caf50;font-style:italic;">✅ Sin pedidos atrasados</p>'
        ESTADO_BG = {'En cotización':'#FFF9C4','No pickeado':'#FFE0B2','No en bulto':'#FFCDD2'}
        t = '<table style="width:100%;border-collapse:collapse;font-size:12px;">'
        t += '<tr>' + ''.join(f'<th style="background:#1B3A6B;color:#fff;padding:8px 12px;text-align:left;">{h}</th>'
                               for h in ['N° Pedido','Cliente','Estado','Umbral','Días']) + '</tr>'
        for r in rows[:max_r]:
            bg = ESTADO_BG.get(r['Estado'],'#fff')
            t += f'''<tr style="background:{bg};">
              <td style="padding:7px 12px;border-bottom:1px solid #eee;font-weight:bold;">{r["N° Pedido"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;">{r["Cliente"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;">{r["Estado"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;color:#888;">{r["Umbral"]}</td>
              <td style="padding:7px 12px;border-bottom:1px solid #eee;text-align:center;font-weight:bold;">{r["Días"]}d</td>
            </tr>'''
        if len(rows) > max_r:
            t += f'<tr><td colspan="5" style="text-align:center;padding:8px;color:#888;font-style:italic;">... y {len(rows)-max_r} más en el Excel adjunto</td></tr>'
        return t + '</table>'

    def seccion(emoji, titulo, n, color, tabla_html):
        return f'''
        <div style="margin-bottom:32px;">
          <h3 style="margin:0 0 12px;color:#1B3A6B;font-size:15px;
                     border-left:4px solid {color};padding-left:12px;">
            {emoji} {titulo}
            <span style="margin-left:10px;background:{color};color:#fff;
                         padding:2px 10px;border-radius:12px;font-size:12px;">{n}</span>
          </h3>
          {tabla_html}
        </div>'''

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f0f2f5;margin:0;padding:24px;">
<div style="max-width:820px;margin:auto;background:#fff;border-radius:10px;
            overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">

  <div style="background:#1B3A6B;padding:28px 32px;">
    <h1 style="color:#fff;margin:0;font-size:22px;">⚠️ Reporte de Alertas — Temponovo</h1>
    <p style="color:#8eadd4;margin:6px 0 0;font-size:13px;">{hoy} · Envío automático Lunes y Jueves</p>
  </div>

  <div style="background:#f8f9fb;padding:20px 32px;border-bottom:1px solid #e8eaed;">
    <div style="display:flex;gap:16px;flex-wrap:wrap;">
      {"".join(f'''<div style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
                  padding:16px 22px;text-align:center;min-width:110px;">
        <div style="font-size:26px;font-weight:bold;color:{vc};">{vv}</div>
        <div style="font-size:11px;color:#888;margin-top:2px;">{vl}</div>
      </div>''' for vv, vc, vl in [
          (total, '#c62828' if total>0 else '#388e3c', 'Total alertas'),
          (len(desc_res), '#D32F2F', 'Descuentos altos'),
          (len(cobr_res), '#E65100', 'Clientes vencidos'),
          (fmt(total_vencido), '#c62828', 'Total vencido &gt;30d'),
          (len(pedidos), '#1565C0', 'Pedidos atrasados'),
      ])}
    </div>
  </div>

  <div style="padding:28px 32px;">
    {seccion('🏷️','Descuentos superiores al 30%', len(desc_res), '#D32F2F', tabla_desc(desc_res))}
    {seccion('💸','Cobranza vencida', len(cobr_res), '#E65100', tabla_cobr(cobr_res))}
    {seccion('📦','Pedidos atrasados por etapa', len(pedidos), '#1565C0', tabla_ped(pedidos))}
  </div>

  <div style="background:#f8f9fb;padding:16px 32px;border-top:1px solid #e8eaed;text-align:center;">
    <p style="margin:0;font-size:11px;color:#aaa;">
      Reporte automático · Temponovo · Datos extraídos de Odoo ERP<br>
      Se adjuntan 3 archivos Excel con el detalle completo
    </p>
  </div>
</div>
</body></html>"""


# ══════════════════════════════════════════════
# ENVÍO EMAIL
# ══════════════════════════════════════════════
def enviar_email(html, excel_desc, excel_cobr, excel_ped, desc_res, cobr_res, pedidos):
    fecha_str = date.today().strftime('%Y%m%d')
    total     = len(desc_res) + len(cobr_res) + len(pedidos)

    msg = MIMEMultipart('mixed')
    msg['From']    = SMTP_USER
    msg['To']      = ', '.join(DESTINATARIOS)
    msg['Subject'] = (
        f"⚠️ Alertas Temponovo — {date.today().strftime('%d/%m/%Y')} "
        f"[{len(desc_res)} desc · {len(cobr_res)} cobr · {len(pedidos)} ped]"
    )
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    for datos, nombre in [
        (excel_desc, f"descuentos_{fecha_str}.xlsx"),
        (excel_cobr, f"cobranza_{fecha_str}.xlsx"),
        (excel_ped,  f"pedidos_atrasados_{fecha_str}.xlsx"),
    ]:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(datos)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{nombre}"')
        msg.attach(part)

    ctx = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
        srv.ehlo(); srv.starttls(context=ctx)
        srv.login(SMTP_USER, SMTP_PASSWORD)
        srv.sendmail(SMTP_USER, DESTINATARIOS, msg.as_bytes())

    print(f"✅ Enviado — {total} alertas ({len(desc_res)} desc · {len(cobr_res)} cobr · {len(pedidos)} ped)")


# ══════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════
def main():
    print(f"\n🚀 Reporte alertas Temponovo — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    uid, models = conectar_odoo()
    print("✅ Conectado a Odoo")

    print("  🏷️  Descuentos...")
    desc_res, desc_det = get_descuentos(models, uid)

    print("  💸 Cobranza...")
    cobr_res, cobr_det = get_cobranza(models, uid)

    print("  📦 Pedidos atrasados...")
    pedidos = get_pedidos_atrasados(models, uid)

    print("  📄 Generando Excel...")
    e_desc = excel_descuentos(desc_det)
    e_cobr = excel_cobranza(cobr_det)
    e_ped  = excel_pedidos(pedidos)

    print("  📧 Enviando email...")
    html = generar_html(desc_res, cobr_res, pedidos)
    enviar_email(html, e_desc, e_cobr, e_ped, desc_res, cobr_res, pedidos)

    print("✅ Proceso completado\n")


if __name__ == "__main__":
    main()
